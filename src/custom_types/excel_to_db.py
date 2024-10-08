from openpyxl import load_workbook, Workbook
from psycopg2 import connect
from custom_types.connection_details import ConnectionDetails
from custom_types.data_type import DataType
from helper_functions.sql_identifier_check import ident_check
from helper_functions.generate_id import generate_id
from datetime import datetime

# When initialised, treat ExcelToDB like a cursor.

class ExcelToDB:
	def __init__(self, file_path : str, db_conn_details : ConnectionDetails):
		self.file_path = file_path
		self.db_conn_details = db_conn_details
		self.column_types : dict[str, DataType] = {}
		try:
			self.__connection__ = connect(
				host=db_conn_details.host,
				port=db_conn_details.port,
				database=db_conn_details.database,
				user=db_conn_details.user,
				password=db_conn_details.password
			)
		except Exception as e:
			raise Exception(f'Error connecting to database\n{e}')

		try:
			self.wb : Workbook = load_workbook(file_path)
		except Exception as e:
			raise Exception(f'Error loading workbook\n{e}')
	
	def swap_active_sheet(self, sheet_name: str):
		'''
		Swaps the active sheet in the workbook to the sheet with the name provided.
		'''

		# Check if the sheet exists.
		if sheet_name not in self.wb.sheetnames:
			raise ValueError(f'Sheet {sheet_name} not found.')
		
		# Set the active sheet to the sheet with the name provided.
		self.wb.active = self.wb[sheet_name]

	def get_column_names(self) -> list[str]:
		'''
		Returns the column names of the active sheet in the workbook.
		'''
		active = self.wb.active

		# I don't know when this wouldn't be the case, but just in case.
		if not active:
			raise TypeError('No active sheet found.')

		# Get the column names.
		columns = active.iter_rows(min_row=1, max_row=1, values_only=True)

		# Convert the columns to a list of strings.
		to_return : list[str] = []

		for column in columns:
			for col in column:
				to_return.append(str(col))
		
		return to_return

	def validate_column_names(self, column_names : list[str], table_name : str) -> bool:
		'''
		Validates that the list of column names exist in the active sheet.
		'''

		# Ident check each identifier to prevent SQL injections

		for column_name in column_names:
			if not ident_check(column_name):
				raise ValueError(f'Column name {column_name} is not a valid identifier.')
		
		if not ident_check(table_name):
			raise ValueError(f'Table name {table_name} is not a valid identifier.')

		# Create a valid SQL representation of the column names.
		names = ', '.join(column_names)
		names_list = list(names)
		names_list.insert(0, '(')
		names_list.append(')')
		names = ''.join(names_list)

		query = f'SELECT exists(SELECT {names} FROM {table_name} LIMIT 1);'

		result = ''
		try:
			cursor = self.__connection__.cursor()
			cursor.execute(query)
			result = cursor.fetchone()
		except:
			return False

		assert result is not None

		return result[0]
		
	
	def insert_column_types(self, data_types : list[DataType]):
		'''
		After the column names have been retrieved, the user can now insert the column types (Assuming use in a CLI context).

		:param data_types: A list of the data types. Can be inserted in any order.
		'''

		column_names = self.get_column_names()

		for data_type in data_types:
			# Should prevent most SQL injection attacks relating to column names.
			if not ident_check(data_type.db_column_name):
				raise ValueError(f'Column name {data_type.db_column_name} is not a valid identifier.')
			
			# Check if the column name exists in the active sheet.
			if data_type.table_column_name not in column_names:
				raise ValueError(f'Column name {data_type.table_column_name} not found in the active sheet.')

			# Add the column type to the dictionary.
			self.column_types[data_type.table_column_name] = data_type
		
		if len(self.column_types) != len(column_names):
			raise ValueError('Not all column types have been inserted.')
	
	def generate_sql(self, table_name : str, randidcol : str | None = None, randidlen : int | None = None) -> list[str]:
		'''
		Generates an SQL query to populate a table.
		'''

		# Get the active sheet.
		active = self.wb.active

		# Denullify the sheet.
		if not active:
			raise TypeError('No active sheet found.')
		
		# Check that the table name is a valid identifier.
		if not ident_check(table_name):
			raise ValueError('Table name is not a valid identifier.')
		
		# Check that the random ID column name is a valid identifier and that the length is valid.
		if randidcol:
			# Check that the column name is a valid identifier.
			if not ident_check(randidcol):
				raise ValueError('Random ID column name is not a valid identifier.')
			
			# Why would anyone need a 255 digit random ID?
			# With 255 digits, you could assign 10 IDs for every grain of sand on Earth.
			if not randidlen or randidlen < 1 or randidlen > 255:
				raise ValueError('Random ID length must be greater than 0 and lower than 255.')
		
		# Generate the SQL statements.
		statements = []

		for row_number in range(2, active.max_row + 1):
			# Key: DB Column Name
			# Value: Value to insert
			to_insert: dict[str, str | int | float | datetime] = {}

			# For each column, get the value in that row.
			for column_number in range(1, active.max_column + 1):
				column_name = str(active.cell(row=1, column=column_number).value)

				# Get data type

				data_type = self.column_types[column_name]

				# Add the value to the dictionary.

				to_insert[data_type.db_column_name] = data_type.parse(str(active.cell(row=row_number, column=column_number).value))
			

			# If a random ID column is to be generated, generate it.
			if randidcol:
				assert randidlen is not None
				to_insert[randidcol] = generate_id(randidlen)
	
			# Generate the SQL statement.
			column_names : list[str] = list(to_insert.keys())

			if not self.validate_column_names(column_names, table_name):
				raise ValueError('Column names do not exist in the table.')

			statement = f'INSERT INTO {table_name} ({', '.join(column_names)}) VALUES ({', '.join(['%s'] * len(column_names))});'

			# Use the connection to "mogrify" the statement.

			statement = self.__connection__.cursor().mogrify(statement, list(to_insert.values())).decode('utf-8')

			# Append the statement to the list of statements.

			statements.append(statement)
		
		self.statements : list[str] = statements
		return statements
	
	def execute_sql(self):
		'''
		Executes the SQL statements generated by generate_sql.
		'''

		# Get statements and ensure they do exist.
		statements = self.statements

		# If not then raise an error.
		if not statements:
			raise ValueError('No statements to execute.')

		# Execute each statement.
		for statement in statements:
			self.__connection__.cursor().execute(statement)
		
		# And commit the changes.
		self.__connection__.commit()

if __name__ == '__main__':
	print('This is part of a library and should not be run directly.')
	exit(1)