from custom_types.excel_to_db import ExcelToDB
from custom_types.connection_details import ConnectionDetails
from custom_types.data_type import DataType
from custom_types.data_type_enum import DataTypeEnum
from dotenv import load_dotenv
from argparse import ArgumentParser
from openpyxl import load_workbook
from os import getenv
from getpass import getpass
from helper_functions.sql_identifier_check import ident_check
import signal
import json

parser = ArgumentParser(description='Excel -> DB: Read an Excel file and insert the data into a PostgreSQL table.')
parser.add_argument('--env-path', '-e', type=str, dest='envpath', help='The path to the .env file to load. Defaults to .env.')
parser.add_argument('--table-name', '-t', type=str, dest='tablename', help='The name of the table to insert the data into.')
parser.add_argument('--json-path', '-j', type=str, dest='jsonpath', help='The path of the JSON file with column data to load.')
parser.add_argument('--assume-yes', '-y', action='store_true', dest='assumeyes', help='Assume yes to prompts.')

group = parser.add_argument_group('XLSX Settings')
group.add_argument('--file-path', '-f', type=str, dest='filepath', help='The path to the Excel file to load.')
group.add_argument('--sheet-name', '-s', type=str, dest='sheetname', help='The name of the sheet to load, if the XLSX file has multiple sheets.')

group = parser.add_argument_group('Random ID Column Generation')
group.add_argument('--rand-col-name', '-c', type=str, dest='randcolname', help='The name of the column to generate ID values for.')
group.add_argument('--rand-col-length', '-l', type=int, dest='randcollength', help='The length of the random IDs to generate.')

args = parser.parse_args()

def signal_handler(sig, frame):
	print('\n\nExiting gracefully.')
	exit(0)

signal.signal(signal.SIGINT, signal_handler)

# If this script is being ran from the commandline, then the file will act as a CLI interface for ExcelToDB class.

if __name__ == '__main__':
	if (args.randcolname and not args.randcollength) or (args.randcollength and not args.randcolname):
		print('--rand-col-name and --rand-col-length must be used together.')
		exit(1)
	
	table_name = ''
	if args.tablename is None:
		while True:
			table_name = input('Enter the name of the table to insert the data into: ')

			if len(table_name) > 63 or not ident_check(table_name):
				print('Table name must be less than 64 characters and must be a valid identifier.')
				continue
			else:
				break
	else:
		table_name = args.tablename

	# Load the Excel file.
	file_path = ''

	while True:
		file_path = args.filepath or input('Enter path to the excel file: ')

		try:
			wb = load_workbook(file_path)
			break
		except:
			print('Failed to load workbook.')
			continue

	# Load the .env file (Hopefully with connection details)
	connection_details = ''
	cursor = None

	# Attempt to load the .env file
	load_dotenv(args.envpath or '.env')

	# Errored once variable to manually prompt user for connection details if connection fails once.
	# Prevents a behaviour where because the .env file exists but the connection details are wrong the system will continually spew errors.

	errored_once = False
	while True:
		# Attempt to grab details
		host = getenv('db_host') if not errored_once else None
		port = getenv('db_port') if not errored_once else None
		user = getenv('db_login') if not errored_once else None
		password = getenv('db_pass') if not errored_once else None
		database = getenv('db_database') if not errored_once else None

		# Seriously, it's wayyy easier to just create a .env file
		if not host:
			host = input('Enter the database host: ')
		if not port:
			port = input('Enter the database port: ')
		if not user:
			user = input('Enter the database username: ')
		if not password:
			password = getpass('Enter the database password (Input hidden): ')
		if not database:
			database = input('Enter the database name: ')
		
		# Attempt to convert port to an integer
		try:
			port = int(port)
			if port < 1 or port > 65535:
				raise ValueError('Port must be between 1 and 65535.')
		except:
			# Or drag the user through the whole process again :)
			print('Port must be an integer.')
			errored_once = True
			continue

		# Create the connection details object
		connection_details = ConnectionDetails(host=host, port=port, user=user, password=password, database=database)

		try:
			cursor = ExcelToDB(file_path, connection_details)
			break
		except:
			# Generally there will only be an error here if the connection details are wrong.
			errored_once = True
			print('Failed to connect to database.')
	
	if len(wb.sheetnames) > 1 and not args.sheetname:
		print('Found multiple sheets. Select a sheet:')
		for sheet in wb.sheetnames:
			print(sheet)
		
		while True:
			sheet_name = input('Enter the name of the sheet to load: ')

			if sheet_name not in wb.sheetnames:
				print('Sheet not found.')
				continue
			else:
				cursor.swap_active_sheet(sheet_name)
				break
	elif len(wb.sheetnames) > 1 and args.sheetname:
		
		if args.sheetname not in wb.sheetnames:
			print('Sheet not found.')
			exit(1)

		cursor.swap_active_sheet(args.sheetname)
	elif len(wb.sheetnames) == 1 and args.sheetname:
		# If only one sheet, and a user specifies an incorrect sheet name, then an error should be thrown.
		if args.sheetname != wb.sheetnames[0]:
			print('Sheet not found.')
			exit(1)

	# Get the column names
	column_names = cursor.get_column_names()

	# Key: Excel Column Name
	# Value: DataType object
	data_types : dict[str, DataType] = {}

	if args.jsonpath:
		print('Attempting to load column data from JSON file.')
		file = ''
		data = ''
		try:
			file = open(args.jsonpath, 'r')
		except:
			print('Failed to open JSON file.')
			exit(1)
		
		try:
			data = json.load(file)
		except:
			print('Invalid JSON file.')
			exit(1)
		
		# JSON Structure:
		# [
		# 	{
		#      "columnName": "Excel Column Name",
		#      "dbColumnName": "Database Column Name",
		#      "columnType": "Data Type"
		#   }
		# ]

		for column in data:
			# Extract the values from the JSON object
			table_column_name : str = column['columnName']
			db_column_name : str = column['dbColumnName']
			column_type : str = column['columnType']

			# Validate that the JSON structure is correct.
			if not table_column_name or not db_column_name or not column_type:
				print('Invalid JSON structure.')
				exit(1)
			
			# Check that the DB column name is a valid identifier
			if not ident_check(db_column_name):
				print(f'Invalid database column name: {db_column_name}')
				exit(1)

			# Check that the table column name exists in the Excel file
			if table_column_name not in column_names:
				print(f'Column name {table_column_name} not found in the Excel file.')
				exit(1)
			
			# Check that the data type is a valid DataTypeEnum
			# If not then assume enum
			if column_type in DataTypeEnum.__members__:
				column_type = DataTypeEnum[column_type].value
			elif not ident_check(column_type):
				print(f'Invalid data type: {column_type}')
				exit(1)
			
			# Construct the DataType object
			data_type_object = DataType(
				table_column_name=table_column_name,
				db_column_name=column['dbColumnName'],
				data_type=column['columnType']
			)

			data_types[table_column_name] = data_type_object
		
		print('Column data loaded from JSON file.')

		# Check that all required column names are present.
		for column_name in column_names:
			if column_name not in list(data_types.keys()):
				print(f'Column name {column_name} not found in JSON file.')
				exit(1)
	else:
		print('You will now be prompted by each column name and asked for a database column name to map this Excel column to.')
		for column_name in column_names:
			print('Excel Column Name: ' + column_name)
			db_column_name = input('Enter the database column name: ')
			data_type : DataTypeEnum | str | None = None
			while True:
				user_input = input('Enter the data type for this column. Enter /? to see a list of data types: ')

				# Help function
				if user_input == '/?':
					print('Data Types:')
					for dt in DataTypeEnum:
						print(dt.value)
					continue
				
				# Attempt to map to a DataTypeEnum
				if user_input in DataTypeEnum.__members__:
					data_type = DataTypeEnum[user_input]
					break
				
				# Otherwise assume the user is attempting to enter a custom data type.
				# Check that the data type is a valid identifier.
				if not ident_check(user_input):
					print('Invalid data type.')
					continue
				else:
					data_type = user_input
					break
			
			data_type_object = DataType(
				table_column_name=column_name,
				db_column_name=db_column_name,
				data_type=data_type
			)

			data_types[column_name] = data_type_object
	
	# Insert the column types into the cursor

	cursor.insert_column_types(list(data_types.values()))

	print('Column types inserted.')

	# Allow for user to generate a random ID column if they didn't specify this in the arguments.

	randidcol = ''
	randidlen = 0
	
	# Will be bypassed if the user uses -y flag.

	if not args.assumeyes and not args.randcolname:
		user_input_randcolname = input('Generate a random ID column? (y/N): ').lower().strip()

		if user_input_randcolname == 'y':
			while True:
				# Column name
				randidcol = input('Enter the name of the random ID column: ')

				# Check that the column name is a valid identifier
				if not ident_check(randidcol):
					print('Invalid random ID column name.')
					continue

				randidlen = input('Enter the length of the random ID column: ')

				# Prevent 3.4 from being a valid length.
				if not randidlen.isdigit():
					print('Random ID length must be an integer.')
					continue

				try:
					randidlen = int(randidlen)
					break
				except:
					print('Random ID length must be an integer.')
					continue
	else:
		randidcol = args.randcolname or None
		randidlen = args.randcollength or None

	# Generate the SQL
	cursor.generate_sql(table_name=table_name, randidcol=randidcol or None, randidlen=randidlen or None)

	print('SQL generated.')

	# Allow the user to check the SQL isn't going to do anything too crazy.
	while True and not args.assumeyes:
		user_input = input('Would you like to see the SQL before execution? (Y/N): ').lower()

		if user_input == 'y':
			print('\n'.join(cursor.statements))
			print('If you do not wish to execute, press CTRL+C/CTRL+Z.')
			continue
		elif user_input == 'n':
			print('Executing.')
			break
		else:
			print('Invalid input.')
			continue

	cursor.execute_sql()

	print('Data inserted. Success.')

	exit(0)